"""Export release-traceable test cases as an Excel workbook."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelExporter:
    def export(self, state: dict) -> bytes:
        release = state.get("release_traceability") or {}
        workbook = Workbook()
        release_sheet = workbook.active
        release_sheet.title = "Release Information"
        release_sheet.append(["Release Information", "Value"])
        release_sheet["A1"].font = release_sheet["B1"].font = Font(bold=True, color="FFFFFF")
        release_sheet["A1"].fill = release_sheet["B1"].fill = PatternFill("solid", fgColor="0078D4")
        fields = [
            ("Jira Key", "jira_key"), ("Jira Type", "jira_type"),
            ("Sprint", "sprint"), ("Sprint Status", "sprint_status"),
            ("CW Release", "cw_release"), ("FA Release Version", "fa_release_version"),
            ("Fix Version", "fix_version"), ("Planned Release Date", "planned_release_date"),
            ("Environment", "environment"), ("Jira URL", "jira_url"),
        ]
        for label, key in fields:
            release_sheet.append([label, release.get(key, "")])
        if release.get("availability_message"):
            release_sheet.append(["Availability", release["availability_message"]])
        release_sheet.column_dimensions["A"].width = 24
        release_sheet.column_dimensions["B"].width = 70

        test_sheet = workbook.create_sheet("Test Cases")
        headers = [
            "TC ID", "Title", "Risk", "AC Reference", "Sprint", "Sprint Status",
            "CW Release", "FA Release Version", "Fix Version", "Environment",
            "Execution Status", "Execution Comments", "Linked Defects", "Evidence",
        ]
        test_sheet.append(headers)
        for cell in test_sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0078D4")
        executions = {
            item.get("tc_id"): item
            for item in ((state.get("test_cycle") or {}).get("executions") or [])
        }
        for test_case in state.get("test_cases") or []:
            execution = executions.get(test_case.get("tc_id"), {})
            test_sheet.append([
                test_case.get("tc_id", ""), test_case.get("title", ""),
                test_case.get("risk_level", ""), test_case.get("ac_ref", ""),
                release.get("sprint", ""), release.get("sprint_status", ""),
                release.get("cw_release", ""), release.get("fa_release_version", ""),
                release.get("fix_version", ""), release.get("environment", ""),
                execution.get("status", "NOT_EXECUTED"),
                execution.get("execution_comments", ""),
                ", ".join(item.get("key", "") for item in execution.get("defects", [])),
                ", ".join(item.get("filename", "") for item in execution.get("evidence", [])),
            ])
        test_sheet.freeze_panes = "A2"

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
