"""Tests for release traceability in generated export formats."""

from __future__ import annotations

import csv
import io
import json

from exports.csv_exporter import CsvExporter
from exports.json_exporter import JsonExporter
from exports.markdown_exporter import MarkdownExporter
from exports.excel_exporter import ExcelExporter
from exports.pdf_exporter import PdfExporter
from openpyxl import load_workbook


def _state() -> dict:
    release = {
        "jira_key": "NGWD6-1",
        "jira_type": "User Story",
        "jira_summary": "Accessible search",
        "jira_url": "https://jira.example/browse/NGWD6-1",
        "sprint": "Sprint 124",
        "sprint_status": "active",
        "cw_release": "CW34",
        "fa_release_version": "FA 2026.08.1",
        "fix_version": "8.14.0",
        "planned_release_date": "2026-08-20",
        "environment": "test",
        "availability_message": "",
    }
    state = {
        "story_id": "NGWD6-1",
        "release_traceability": release,
        "structured_requirement": {"summary": "Accessible search"},
        "test_cases": [{
            "tc_id": "TC-001",
            "title": "Keyboard focus",
            "steps": [],
            "tags": [],
            "browser_coverage": [],
            "mobile_coverage": [],
            "evidence_required": ["Screenshot"],
        }],
        "test_data": {},
        "coverage_map": {},
        "review_report": {},
        "final_report": {},
        "rtm": {"rows": [{"row_id": "RTM-001", "story_id": "NGWD6-1", "tc_id": "TC-001"}]},
    }
    state["test_cycle"] = {
        "cycle_id": "TCY-1",
        "name": "Sprint 124 - NGWD6-1 Validation",
        "environment": "test",
        "execution_type": "Functional Testing",
        "executions": [{
            "tc_id": "TC-001",
            "status": "FAIL",
            "execution_comments": "Focus is obscured",
            "defects": [{"key": "BUG-456"}],
            "evidence": [{"filename": "focus.png"}],
        }],
        "metrics": {
            "total": 1, "executed": 1, "passed": 0, "failed": 1,
            "blocked": 0, "not_executed": 0, "execution_progress_pct": 100.0,
            "pass_rate_pct": 0.0, "defect_count": 1,
        },
        "readiness_summary": {
            "requirement_coverage_pct": 100.0,
            "risk_assessment": "HIGH",
            "release_recommendation": "NO GO",
            "linked_defects": ["BUG-456"],
        },
    }
    return state


def test_markdown_starts_with_release_information_and_repeats_per_test():
    output = MarkdownExporter().export(_state())

    assert output.startswith("# Release Information")
    assert "**CW Release:** CW34" in output
    assert output.count("**FA Release Version:** FA 2026.08.1") >= 2
    assert "**Test Evidence Traceability:**" in output
    assert "**Execution Status:** FAIL" in output
    assert "**Release Recommendation:** NO GO" in output


def test_json_includes_release_information_per_test():
    payload = json.loads(JsonExporter().export(_state()))

    assert payload["release_information"]["sprint"] == "Sprint 124"
    assert payload["tests"][0]["xray_metadata"]["release_traceability"]["cw_release"] == "CW34"
    assert payload["test_cycle"]["executions"][0]["defects"][0]["key"] == "BUG-456"


def test_csv_includes_release_information_on_every_row():
    rows = list(csv.DictReader(io.StringIO(CsvExporter().export(_state()))))

    assert rows[0]["Sprint"] == "Sprint 124"
    assert rows[0]["CW Release"] == "CW34"
    assert rows[0]["FA Release Version"] == "FA 2026.08.1"
    assert rows[0]["Execution Status"] == "FAIL"
    assert rows[0]["Linked Defects"] == "BUG-456"


def test_excel_starts_with_release_sheet_and_repeats_test_traceability():
    workbook = load_workbook(io.BytesIO(ExcelExporter().export(_state())))

    assert workbook.sheetnames[0] == "Release Information"
    assert workbook["Release Information"]["B4"].value == "Sprint 124"
    assert workbook["Test Cases"]["E2"].value == "Sprint 124"
    assert workbook["Test Cases"]["K2"].value == "FAIL"


def test_pdf_contains_a_valid_report():
    output = PdfExporter().export(_state())

    assert output.startswith(b"%PDF")
    assert len(output) > 1000
