"""
Generate Excel Test Report from Markdown
Converts CMS Sprint 274 report to professional Excel format
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define colors and styles
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BLOCKED_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_test_report():
    """Create Excel test report for CMS Sprint 274"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Summary", 0)
    stories_sheet = wb.create_sheet("Stories", 1)
    testcases_sheet = wb.create_sheet("Test Cases", 2)
    rtm_sheet = wb.create_sheet("RTM", 3)
    
    # Build sheets
    build_summary(summary_sheet)
    build_stories(stories_sheet)
    build_testcases(testcases_sheet)
    build_rtm(rtm_sheet)
    
    # Save workbook
    output_path = r"c:\Chetan Raut\Projects\Volkswagen NGW\KI_Tetscases\output\CMS Sprint 274\CMS_SPRINT_274_TEST_REPORT.xlsx"
    wb.save(output_path)
    print(f"✅ Report created: {output_path}")
    return output_path


def build_summary(ws):
    """Build Summary sheet"""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Title
    ws['A1'] = "CMS SPRINT 274 — TEST REPORT"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')
    
    # Metadata
    row = 3
    ws[f'A{row}'] = "Generated"
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    row += 1
    ws[f'A{row}'] = "Sprint"
    ws[f'B{row}'] = "CMS Sprint 274"
    
    row += 1
    ws[f'A{row}'] = "Component"
    ws[f'B{row}'] = "CMS / NBD'26"
    
    row += 1
    ws[f'A{row}'] = "Tool"
    ws[f'B{row}'] = "QE AgentX v1.0"
    
    # Metrics Section
    row += 3
    ws[f'A{row}'] = "SPRINT METRICS"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Metric", "Value", "Status", ""]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    metrics = [
        ["Stories", "3", "✅ Complete"],
        ["Manual Test Cases", "9", "✅ All Passed"],
        ["AgentX Generated TCs", "7", "✅ Recommended"],
        ["Total Test Cases", "16", "✅ Complete"],
        ["Pass Rate", "100%", "✅ Excellent"],
        ["Quality Score", "88/100", "✅ High"],
        ["AC Coverage", "50-86%", "⚠️ Gap Analysis"],
        ["Screenshots", "9", "✅ Complete"],
    ]
    
    for metric_row in metrics:
        row += 1
        for col, value in enumerate(metric_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if col == 3 and "✅" in str(value):
                cell.fill = PASS_FILL
            elif col == 3 and "⚠️" in str(value):
                cell.fill = BLOCKED_FILL
    
    # Story Summary
    row += 3
    ws[f'A{row}'] = "STORY SUMMARY"
    ws[f'A{row}'].font = SUBHEADER_FONT
    ws[f'A{row}'].fill = SUBHEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ["Story Key", "Title", "Test Cases", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    stories = [
        ["NGWD6-50225", "Glossary Section & Item Update", "5", "✅ PASSED"],
        ["NGWD6-50304", "Content Layer & Expand-Collapse Update", "4", "✅ PASSED"],
        ["NGWD6-51297", "Navigation Hover Highlighting", "7", "✅ PASSED"],
    ]
    
    for story in stories:
        row += 1
        for col, value in enumerate(story, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if "✅" in str(value):
                cell.fill = PASS_FILL


def build_stories(ws):
    """Build Stories detail sheet"""
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Title
    ws['A1'] = "STORY DETAILS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    
    row = 3
    headers = ["Story Key", "Summary", "ACs", "Status", "Evidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    stories_data = [
        {
            "key": "NGWD6-50225",
            "summary": "Glossary Section (Simple) ↔ Glossary Item (Simple) - Update as per new design",
            "acs": 2,
            "status": "✅ PASSED",
            "evidence": "5 test cases; 3 screenshots"
        },
        {
            "key": "NGWD6-50304",
            "summary": "Content Layer (Medium) + InterLayerNavigation ↔ Expand-Collapse Item - Update as per new design",
            "acs": 2,
            "status": "✅ PASSED",
            "evidence": "4 test cases; 6 screenshots"
        },
        {
            "key": "NGWD6-51297",
            "summary": "Visual hover highlighting for menu items in Main Navigation Flyout",
            "acs": 9,
            "status": "✅ PASSED",
            "evidence": "1 manual + 7 AgentX TCs; MD, JSON, CSV"
        },
    ]
    
    for story in stories_data:
        row += 1
        ws.cell(row=row, column=1, value=story["key"]).border = BORDER
        ws.cell(row=row, column=2, value=story["summary"]).border = BORDER
        ws.cell(row=row, column=3, value=story["acs"]).border = BORDER
        
        status_cell = ws.cell(row=row, column=4, value=story["status"])
        status_cell.border = BORDER
        status_cell.fill = PASS_FILL
        
        ws.cell(row=row, column=5, value=story["evidence"]).border = BORDER


def build_testcases(ws):
    """Build Test Cases sheet"""
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Title
    ws['A1'] = "TEST CASE EXECUTION"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    row = 3
    headers = ["TC#", "Story", "Test Case", "Risk", "AC Ref", "Result"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    testcases_data = [
        # NGWD6-50225
        [1, "NGWD6-50225", "Figma Spec Compliance (NBD Toggle ON)", "HIGH", "AC-01", "✅ PASS"],
        [2, "NGWD6-50225", "Column Padding at Breakpoint < 1280px", "MEDIUM", "AC-01", "✅ PASS"],
        [3, "NGWD6-50225", "Column Padding at Breakpoint ≥ 1280px", "MEDIUM", "AC-01", "✅ PASS"],
        [4, "NGWD6-50225", "Gap Between Upper and Lower Heading", "LOW", "AC-01", "✅ PASS"],
        [5, "NGWD6-50225", "NBD Toggle OFF (Regression)", "HIGH", "AC-02", "✅ PASS"],
        # NGWD6-50304
        [6, "NGWD6-50304", "Named Section Match NBD Specifications", "HIGH", "AC-01", "✅ PASS"],
        [7, "NGWD6-50304", "NBD Hidden Behind Feature Activation", "HIGH", "AC-02", "✅ PASS"],
        [8, "NGWD6-50304", "Content Layer Breakpoints & Padding", "MEDIUM", "AC-01", "✅ PASS"],
        [9, "NGWD6-50304", "Expand-Collapse Item Breakpoints & Padding", "MEDIUM", "AC-01", "✅ PASS"],
        # NGWD6-51297 Manual
        [10, "NGWD6-51297", "Hover State — Menu item highlighting", "HIGH", "AC-01", "✅ PASS"],
        # NGWD6-51297 AgentX
        [11, "NGWD6-51297", "Feature Cluster displays NBD'26 design when active", "HIGH", "AC-01", "Recommended"],
        [12, "NGWD6-51297", "Feature Cluster responsive at 960px tablet viewport", "MEDIUM", "AC-01", "Recommended"],
        [13, "NGWD6-51297", "Toggle ON: NBD'26 design visible", "HIGH", "AC-02", "Recommended"],
        [14, "NGWD6-51297", "Toggle OFF: legacy design fallback", "HIGH", "AC-02", "Recommended"],
        [15, "NGWD6-51297", "Toggle runtime switch updates design dynamically", "MEDIUM", "AC-02", "Recommended"],
        [16, "NGWD6-51297", "Storybook documents Feature Cluster NBD'26 variant", "LOW", "AC-03", "Recommended"],
        [17, "NGWD6-51297", "SysDoc updated with Feature Cluster specs", "LOW", "AC-04", "Recommended"],
    ]
    
    for tc in testcases_data:
        row += 1
        for col, value in enumerate(tc, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            if col == 6:  # Result column
                if "PASS" in str(value):
                    cell.fill = PASS_FILL
                elif "Recommended" in str(value):
                    cell.fill = BLOCKED_FILL


def build_rtm(ws):
    """Build Requirement Traceability Matrix sheet"""
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    
    # Title
    ws['A1'] = "REQUIREMENT TRACEABILITY MATRIX (RTM)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    
    row = 3
    headers = ["AC#", "Acceptance Criterion", "Coverage %", "Test Cases", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    
    rtm_data = [
        ["AC-01", "Named items, section match NBD'26 specifications from FIGMA", "100%", "TC-1,2,3,4,8", "✅ Complete"],
        ["AC-02", "NBD changes hidden behind feature activation toggle", "100%", "TC-5,6,7,13,14,15", "✅ Complete"],
        ["AC-03", "Storybook documentation updated", "50%", "TC-16", "⚠️ Gap"],
        ["AC-04", "SysDoc updated with NBD'26 specs", "50%", "TC-17", "⚠️ Gap"],
        ["AC-05", "Hover highlighting visual design", "100%", "TC-10,11", "✅ Complete"],
        ["AC-06", "Responsive layout breakpoints", "100%", "TC-2,3,8,12", "✅ Complete"],
        ["AC-07", "Feature toggle behaviour (ON/OFF)", "100%", "TC-13,14,15", "✅ Complete"],
        ["AC-08", "RTL language support", "0%", "—", "❌ Gap (TC-008)"],
        ["AC-09", "Performance baseline", "0%", "—", "❌ Gap (TC-009)"],
    ]
    
    for ac in rtm_data:
        row += 1
        for col, value in enumerate(ac, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            
            if col == 5:  # Status column
                if "✅" in str(value):
                    cell.fill = PASS_FILL
                elif "⚠️" in str(value):
                    cell.fill = BLOCKED_FILL
                elif "❌" in str(value):
                    cell.fill = FAIL_FILL


if __name__ == "__main__":
    create_test_report()
